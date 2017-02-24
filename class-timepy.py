import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

WEEK_DAY = {"Lunedì": "A", "Martedì": "B", "Mercoledì": "C", "Giovedì": "D", "Venerdì": "E"}
LESS_TO_COL = {}
COLOR = ["FFE699", "66CDAA", "DEB887", "DDA0DD", "D3D3D3"]


def input_string(path):
    txt = open(path, "r").read()
    if txt:
        return txt
    else:
        input_string()


def lesson_info(string, subject):  # GET LESSON INFO FROM A LINE

    AULAPREFIX = "aula"
    BUILDPREFIX = "Edificio"
    strings = string.split()

    aula, building = 0, 0

    for i, e in enumerate(strings):
        if e == AULAPREFIX:
            aula = strings[i + 1]
        if e == BUILDPREFIX:
            building = strings[i + 1]

    print(WEEK_DAY[strings[0]], strings[2], strings[4], aula, building, subject)
    return (WEEK_DAY[strings[0]], strings[2], strings[4], aula, building, subject)
    # Return DAY STARTT ENDT AULA BUILDING


def check_time_line(line):
    if len(line) > 1:
        return line.split()[0] in WEEK_DAY


def extrapolate(text):  # From the pasted text extrapolate lessons and time

    lessons = []
    subjects = re.split('\d{6} - ', text)

    for subject in subjects:
        raw_subj_name = re.search("\A.*", subject)  # Name of the Subject
        if raw_subj_name:
            subj_name = (raw_subj_name.group(0)).split('(')[0].strip()

        lines = str.splitlines(subject)

        for line in lines:
            if check_time_line(line):
                lessons.append(lesson_info(line, subj_name))

    return lessons


def hour_to_cell(hour):

    return str(int(hour[:2]) - 7)


def number_to_col(numb):

    NUMB_TO_COL = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5, "G": 6, "H": 7}
    return int(NUMB_TO_COL[numb]) + 1


def lesson_to_color(lesson):

    if lesson not in LESS_TO_COL:
        LESS_TO_COL[lesson] = COLOR[0]
        COLOR.remove(COLOR[0])
    return LESS_TO_COL[lesson]


def create_table(lessons):

    wb = Workbook()

    ws = wb.active

    for lesson in lessons:
        start = hour_to_cell(lesson[1])
        end = hour_to_cell(lesson[2])

        first_cell = lesson[0] + start

        ws[first_cell] = str(lesson[4]) + ":" + str(lesson[3]) + " - " + lesson[5]  # WRITE THE INFO IN THE FIRST CELL
        ws[first_cell].alignment = Alignment(horizontal='center')
        ws[first_cell].alignment = Alignment(vertical='center')
        ws[first_cell].fill = PatternFill(start_color=lesson_to_color(lesson[5]),
                                          end_color=lesson_to_color(lesson[5]),
                                          fill_type='solid')

        ws.merge_cells(start_row=int(start), start_column=number_to_col(lesson[0]), end_row=int(end) - 1,
                       end_column=number_to_col(lesson[0]))  # MERGE ALL THE CELL

        ws.column_dimensions[str(lesson[0])].width = 20

    wb.save("Test.xlsx")


def main():
    text = input_string("txt.txt")
    print(text)
    lessons = extrapolate(text)
    # print(lessons) # DEBUG
    create_table(lessons)


if __name__ == '__main__':
    main()
